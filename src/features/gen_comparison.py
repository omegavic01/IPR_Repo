"""
Objective:
    After manually updating the below fields needed for proper comparison.
The script will then perform a diff between the datasets.

Goal:
    The request was to take a list of subnets from a division and perform a
diff between those subnets and what is in the DDI database.
"""
import os
import logging
from xlrd import open_workbook
import openpyxl
from dotenv import find_dotenv, load_dotenv


def _write_output_to_master(diff_list, mod_file, output_file):
    """
        Write Output
    """
    w_b = openpyxl.load_workbook(filename=mod_file)
    w_s = w_b.create_sheet('Not in DDI', 2)
    w_s.title = 'Not in DDI'
    for row_indx, stuff in enumerate(diff_list):
        for col_indx, item in enumerate(stuff):
            w_s.cell(row=row_indx+1, column=col_indx + 1, value=item)
    w_b.save(output_file)


def main_phase_2(ipr_src_file, ipr_src_mod_file, output_file, logger):
    """

    :param ipr_src_file:
    :param ipr_src_mod_file:
    :param output_file:
    :param logger:
    """
    logger.info('Loading Data from .xlsx files')
    # Original Dataset
    ipr = open_workbook(ipr_src_file)
    ipr_sheet = ipr.sheet_by_index(1)
    # Modified Dataset
    ipr_mod = open_workbook(ipr_src_mod_file)
    ipr_mod_sheet = ipr_mod.sheet_by_index(0)

    # Add Datasets to a python list.
    updatelist = []
    updatelist.append(HEADER_ROW[0:17])

    def gather_data(ipr_w_s, ipr_mod_w_s):
        raw_data = []
        mod_data = []
        for idx in range(ipr_sheet.nrows):
            ipr_tup = (ipr_w_s.row_values(idx)[1].strip(),
                       ipr_w_s.row_values(idx)[8].strip())
            raw_data.append(ipr_tup)
        for idx in range(ipr_mod_w_s.nrows):
            mod_tup = (ipr_mod_w_s.row_values(idx)[1].strip(),
                       ipr_mod_w_s.row_values(idx)[8].strip())
            mod_data.append(mod_tup)
        return raw_data, mod_data

    ipr_data, ipr_mod_data = gather_data(ipr_sheet, ipr_mod_sheet)

    logger.info('Building Diff List from modified file.')
    # Build Diff List for processing.
    for enum, ipr_mod_row in enumerate(ipr_mod_data):
        if enum == 0:
            continue
        if ipr_mod_row not in ipr_data:
            updatelist.append(ipr_mod_sheet.row_values(enum)[0:14])
            continue
    logger.info('Writing output to data\\processed folder.')
    _write_output_to_master(updatelist, ipr_src_mod_file, output_file)
    logger.info('Script Complete')


def main():
    """
        Script that takes in two .xlsx files in IPR format.  Performs a diff
    against them and generates output to a separate sheet within mod file.
    """
    logger = logging.getLogger('gen_comparison.py')
    logger.info('Beginning of Script')
    logger.info('Building Paths and Filenames')
    # Build path's.
    interim_data_path = os.path.join(PROJECT_DIR, 'data', 'interim')
    processed_data_path = os.path.join(PROJECT_DIR, 'data', 'processed')

    # Join file names to path's.
    ipr_file = os.path.join(interim_data_path,
                            'DDI_to_IPR.xlsx')
    mod_file_name = 'Copy of Consolidated and formatted (+1).xlsx'
    ipr_mod_file = os.path.join(interim_data_path,
                                mod_file_name)
    out_file = os.path.join(processed_data_path,
                            'With Not in DDI Data ' + mod_file_name)
    main_phase_2(ipr_file, ipr_mod_file, out_file, logger)


if __name__ == '__main__':
    # getting root directory
    PROJECT_DIR = os.path.join(os.path.dirname(__file__), os.pardir, os.pardir)

    # setup logger
    LOG_FMT = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    logging.basicConfig(level=logging.INFO, format=LOG_FMT)

    # find .env automatically by walking up directories until it's found
    DOTENV_PATH = find_dotenv()

    # load up the entries as environment variables
    load_dotenv(DOTENV_PATH)

    # Headerrow for IPR Output
    HEADER_ROW = os.environ.get("IPR_HEADER_ROW").split(',')

    main()
