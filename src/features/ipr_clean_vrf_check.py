"""
This script as part of the IPR initiative takes in the pickled data compiled
from the master audit script.  Compares the VRF's to the other VRF's in order
to find out which VRF conflicts with which VRF.  A seperate function is used
to identify any VRF's that have no conflicts.  The data compiled from these
functions are then added to the DDI_to_IPR.xlsx spreadsheet located in reports.
"""
import logging
import os
import pickle
import openpyxl


def _write_out_conflict_vrf(o_c_data, source_wb):
    w_b = openpyxl.load_workbook(filename=source_wb)
    wb_sheet_names = w_b.sheetnames
    wb_sheet_names.append('Filt-Conflicting-VRF')
    wb_sheet_names.sort(reverse=True)
    sheet_index = wb_sheet_names.index('Filt-Conflicting-VRF')
    w_s = w_b.create_sheet(wb_sheet_names[sheet_index], sheet_index)
    row = 0
    for enum, dic in enumerate(o_c_data):
        key = ''
        for item in dic.keys():
            key = item
        if enum == 0 and dic[key]:
            w_s.cell(row=row + 1, column=1, value='Conflict-VRF 3 digit value')
            row += 1
            w_s.cell(row=row + 1, column=1, value=str(key))
            w_s.cell(row=row + 1, column=2, value=','.join(list(set(
                dic[key]))))
            row += 1
        else:
            if dic[key]:
                w_s.cell(row=row + 1, column=1, value=str(key))
                w_s.cell(row=row + 1, column=2, value=','.join(list(set(
                    dic[key]))))
                row += 1
    w_b.save(source_wb)


def _write_out_clear_vrf(vrf_data, source_wb):
    w_b = openpyxl.load_workbook(filename=source_wb)
    wb_sheet_names = w_b.sheetnames
    wb_sheet_names.append('Filt-Clean-VRF')
    wb_sheet_names.sort(reverse=True)
    sheet_index = wb_sheet_names.index('Filt-Clean-VRF')
    w_s = w_b.create_sheet(wb_sheet_names[sheet_index], sheet_index)
    for enum, i in enumerate(vrf_data):
        if enum == 0:
            w_s.cell(row=enum + 1, column=1, value='Clean-VRF 3 digit value')
            w_s.cell(row=enum + 2, column=1, value=int(i[0]))
        else:
            w_s.cell(row=enum + 2, column=1, value=int(i[0]))
    w_b.save(source_wb)


def _check_vrf_record_uncontested_vrfs(vrf_o_c_dict, vrf_idx):
    return_list = []
    for key in vrf_o_c_dict.keys():
        temp_dict = {}
        temp_dict[key] = []
        for o_c in vrf_o_c_dict[key]:
            if o_c in vrf_idx and key not in vrf_idx[o_c][15]:
                temp_dict[key].append(vrf_idx[o_c][15])
        return_list.append(temp_dict)
    return return_list


def _check_vrf_against_entire_db(vrf_dict):
    clean_vrf = []
    for key in vrf_dict.keys():
        temp_vrf = []
        for i in vrf_dict[key]:
            if 'NO' in i[25] or 'NO' in i[26]:
                temp_vrf.append('NO')
                break
            else:
                temp_vrf.append(key)
        if not temp_vrf or 'NO' in temp_vrf:
            continue
        else:
            clean_vrf.append(list(set(temp_vrf)))
    return clean_vrf


def _compiling_data(data):
    vrf_idx_data_dict = {}
    vrf_dict = {}
    vrf_o_and_c_dict = {}
    for i in data:
        if i[15].startswith('00'):
            vrf_idx_data_dict[i[22]] = i

            def _get_vrf_o_c_dict(i, vrf_o_c_dict):
                if i[15].split('-')[0] not in vrf_o_c_dict:
                    vrf_o_c_dict[i[15].split('-')[0]] = []
                if i[15].split('-')[0] in vrf_o_c_dict:
                    if not i[23]:
                        pass
                    elif i[23] and isinstance(i[23], str):
                        vrf_o_c_dict[i[15].split('-')[0]] += \
                            list(map(int, i[23].split(',')))
                    elif i[23] and isinstance(i[23], int):
                        vrf_o_c_dict[i[15].split('-')[0]].append(i[23])
                    else:
                        if i[23] and isinstance(i[23], str):
                            vrf_o_c_dict[i[15].split('-')[0]] += \
                                list(map(int, i[23].split(',')))
                        elif i[23] and isinstance(i[23], int):
                            vrf_o_c_dict[i[15].split('-')[0]].append(i[23])
                    if not i[24]:
                        pass
                    elif i[24] and isinstance(i[24], str):
                        vrf_o_c_dict[i[15].split('-')[0]] += \
                            list(map(int, i[24].split(',')))
                    elif i[24] and isinstance(i[24], int):
                        vrf_o_c_dict[i[15].split('-')[0]].append(i[24])
                    else:
                        if i[24] and isinstance(i[24], str):
                            vrf_o_c_dict[i[15].split('-')[0]] += \
                                list(map(int, i[24].split(',')))
                        elif i[24] and isinstance(i[24], int):
                            vrf_o_c_dict[i[15].split('-')[0]].append(i[24])
            _get_vrf_o_c_dict(i, vrf_o_and_c_dict)
            if i[15].split('-')[0] not in vrf_dict:
                vrf_dict[i[15].split('-')[0]] = [i]
            else:
                vrf_dict[i[15].split('-')[0]].append(i)
    return vrf_idx_data_dict, vrf_dict, vrf_o_and_c_dict


def _load_pickle_data(file, logger):
    data = []
    try:
        data = pickle.load(open(file, 'rb'))
    except FileNotFoundError:
        logger.error('Pickled data could not be found.  Make sure audit '
                     'process has been completed.  File should be located in '
                     'the /data/processed directory.')
        exit()
    return data


def main():
    """
    Here is where the file paths and filenames are assigned.  Once completed
    the calls to the functions to build the information are made.

    :return:
    """
    logger = logging.getLogger('ipr_clean_vrf_check.py')
    logger.info('Beginning of Script')
    logger.info('Building Paths and Filenames')
    # Build path's.
    processed_data_path = os.path.join(PROJECT_DIR, 'data', 'processed')
    reports_data_path = os.path.join(PROJECT_DIR, 'reports')

    # Join file names to path's.
    source_file = os.path.join(processed_data_path, 'ddi_to_ipr.pkl')
    master_file = os.path.join(reports_data_path, 'DDI_to_IPR.xlsx')

    logger.info('Loading Pickled Data')
    ipr_data = _load_pickle_data(source_file, logger)
    logger.info('Completed')

    logger.info('Compiling data.')
    vrf_idx, vrf_dict, vrf_o_c_dict = _compiling_data(ipr_data)

    logger.info('Checking for VRF with no conflicts or overlaps')
    clear_vrf = _check_vrf_against_entire_db(vrf_dict)
    logger.info('Checking VRFs to compile vrf to clean vrf comparison.')
    vrf_summaries = _check_vrf_record_uncontested_vrfs(vrf_o_c_dict, vrf_idx)

    logger.info('Writing to DDI_to_IPR.xlsx')
    _write_out_clear_vrf(clear_vrf, master_file)
    _write_out_conflict_vrf(vrf_summaries, master_file)
    logger.info('Script Complete')


if __name__ == '__main__':
    # getting root directory
    PROJECT_DIR = os.path.join(os.path.dirname(__file__), os.pardir, os.pardir)

    # setup logger
    LOG_FMT = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    logging.basicConfig(level=logging.INFO, format=LOG_FMT)

    main()
