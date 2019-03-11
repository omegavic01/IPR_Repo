"""Initial Documentation

The hope is to take a template file with builtin formulas within
the .xlsx file.  Then from another .xlsx file copy the data from this sheet
and have it updated the second sheet of the first .xlsx file.  To then have
the original .xlsx file update sheet one formulas with the second sheets
data.  Here goes!"""
import logging
import os
import openpyxl
from openpyxl.styles import Alignment


def copy_data_over(source, template, final):
    """Reference to the site that was used for the below for loop:

    URL:
    https://stackoverflow.com/questions/44593705/how-to-copy-over-an-excel-
    sheet-to-another-workbook-in-python
    """
    source_wb = openpyxl.load_workbook(filename=source)
    source_ws = source_wb['MASTER']
    template_wb = openpyxl.load_workbook(filename=template)
    template_ws = template_wb.worksheets[1]
    for row in source_ws:
        for cell in row:
            template_ws[cell.coordinate].value = cell.value
    max_row = template_ws.max_row
    for row in template_ws.iter_rows(min_row=2, max_row=max_row,
                                     min_col=24, max_col=25):
        for cell in row:
            cell.alignment = Alignment(horizontal='left')
    template_wb.save(final)


def main():
    """Straight forward copying and pasting of data from one .xlsx file
    to another .xlsx file."""
    logger = logging.getLogger('ipr_report_percent.py')
    logger.info('Beginning of Script')
    logger.info('Building Paths and Filenames')
    # Build path's.
    interim_data_path = os.path.join(PROJECT_DIR, 'data', 'interim')
    reports_data_path = os.path.join(PROJECT_DIR, 'reports')

    # Join file names to path's.
    source_file = os.path.join(reports_data_path, 'DDI_to_IPR.xlsx')
    mtr_prcnt_tmplt = os.path.join(interim_data_path,
                                   'MASTER - Report by percent template.xlsx')
    mtr_prcnt_file = os.path.join(reports_data_path,
                                  'MASTER - Report by percent.xlsx')

    logger.info('Writing Data Over')
    copy_data_over(source_file, mtr_prcnt_tmplt, mtr_prcnt_file)
    logger.info('Script Complete')


if __name__ == '__main__':
    # getting root directory
    PROJECT_DIR = os.path.join(os.path.dirname(__file__), os.pardir, os.pardir)

    # setup logger
    LOG_FMT = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    logging.basicConfig(level=logging.INFO, format=LOG_FMT)

    main()
