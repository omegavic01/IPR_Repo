"""
This script is intended to take in the DDI data generated from the
ipr_initial_data_gathering.py script.  Twist, mash, split, and sort the
data.

DataSets:
   ddi_workbook.xls -- input file
   DDI_IPR_Unsorted.xlsx -- temp output file
   DDI_IPR_Sorted.xlsx -- output file
"""
import os
import logging
import operator
import pickle
from xlrd import open_workbook
from ipaddr import IPv4Network
from openpyxl import Workbook
from openpyxl.styles import Alignment
from dotenv import find_dotenv, load_dotenv


def _sorting_data(interim_sorted_file, interim_unsorted_file):
    """
    Takes in the converted data from sort's all of the networks within
    each sheet within the work book.
    """
    input_wb = open_workbook(interim_unsorted_file)
    sheet_names = input_wb.sheet_names()
    output_wb = Workbook()
    for enum, sheet in enumerate(sheet_names):
        input_ws = input_wb.sheet_by_name(sheet)
        sorting_stuff = []  # Taking Unsorted data then sorting.
        for i in range(input_ws.nrows):
            if i == 0:
                continue
            sorting_stuff.append(input_ws.row_values(i))
        sorting_stuff = sorted(sorting_stuff,
                               key=operator.itemgetter(17, 18, 19, 20, 21))

        # Creating new spreadsheet with sorted data.

        def _write_final_output(sorted_stuff, sheet_name, idx):
            output_ws = output_wb.create_sheet(sheet_name, idx)
            output_ws.title = sheet_name
            row = 0
            for index, item in enumerate(HEADER_ROW):
                output_ws.cell(row=row + 1, column=index + 1, value=item)
            for stuff in sorted_stuff:
                row = row + 1
                for index, items in enumerate(stuff):
                    output_ws.cell(row=row + 1, column=index + 1, value=items)
        _write_final_output(sorting_stuff, sheet, enum)
    if 'Sheet' in output_wb.sheetnames:
        std = output_wb['Sheet']
        output_wb.remove(std)
    output_wb.save(interim_sorted_file)


def _write_output_to_master(idx, ddi_dic, path):
    """
    Write out the rows in IPR expected format.

    Output File:
        output_file - DDI_to_IPR_Unsorted.xlsx
    """
    sheet_name_list = []
    for i in ddi_dic.keys():
        sheet_name_list.append(i)
    sheet_name_list.sort()
    sheet_name_list.reverse()

    # Buildout workbook!
    work_book = Workbook()
    for enum, sheet_name in enumerate(sheet_name_list):
        if not ddi_dic[sheet_name]:
            continue
        w_s = work_book.create_sheet(sheet_name, enum)  # work_sheet
        w_s.title = sheet_name
        row = 1
        for index, item in enumerate(HEADER_ROW):
            w_s.cell(row=row, column=index + 1, value=item)
        for bit in ddi_dic[sheet_name]:
            row = row + 1
            w_s.cell(row=row, column=2, value=bit[3])  # Cidr
            w_s.cell(row=row, column=3, value=bit[idx['Region_List']])
            w_s.cell(row=row, column=4, value=bit[idx['CO']])
            w_s.cell(row=row, column=5, value=bit[idx['City']])
            w_s.cell(row=row, column=6, value=bit[idx['Address']])
            w_s.cell(row=row, column=7, value=bit[idx['Site']])
            w_s.cell(row=row, column=8, value=bit[idx['Datacenter']])
            w_s.cell(row=row, column=9, value=bit[idx['Div']])
            w_s.cell(row=row, column=10, value=bit[idx['Req Email']])
            w_s.cell(row=row, column=11, value=bit[idx['Agency']])
            w_s.cell(row=row, column=13, value=bit[5])  # comment
            w_s.cell(row=row, column=12, value=bit[
                idx['VLAN Description']])
            w_s.cell(row=row, column=14, value=bit[
                idx['Interface Name']])
            w_s.cell(row=row, column=15, value=bit[0])  # ddi_type
            w_s.cell(row=row, column=16, value=bit[4])  # ddi_view
            w_s.cell(row=row, column=17, value=bit[idx['IPR D']])
            mycell = w_s.cell(row=row, column=18)
            mycell.alignment = Alignment(horizontal='left')
            mycell.value = int(bit[3].split('.')[0])  # 1st octet
            mycell = w_s.cell(row=row, column=19)
            mycell.alignment = Alignment(horizontal='left')
            mycell.value = int(bit[3].split('.')[1])  # 2nd octet
            mycell = w_s.cell(row=row, column=20)
            mycell.alignment = Alignment(horizontal='left')
            mycell.value = int(bit[3].split('.')[2])  # 3rd octet
            mycell = w_s.cell(row=row, column=21)
            mycell.alignment = Alignment(horizontal='left')
            mycell.value = int(bit[3].split('.')[3].split('/')[0])  # 4th octet
            mycell = w_s.cell(row=row, column=22)
            mycell.alignment = Alignment(horizontal='left')
            mycell.value = int(bit[3].split('/')[1])  # cidr for network addr.
    if 'Sheet' in work_book.sheetnames:
        std = work_book['Sheet']
        work_book.remove(std)
    work_book.save(path)


def _filter_data(file):
    rddi = open_workbook(file)
    rddifirst_sheet = rddi.sheet_by_index(0)

    omc_it_parent_list = list({
        "CDS - Guest Range 1",
        "PROD WEST",
        "MGMT WEST",
        "VOICE WEST",
        "NONAGENCY West",
        "NONAGENCY EAST",
        "AGENCY WEST",
        "Agency West",
        "PROD EAST",
        "MGMT EAST",
        "VOICE EAST",
        "HUB EAST",
        "Agency East Space",
        "Administrative Container",
        "AGENCY EMEA SPACE",
        "EMEA - Users @ Bankside",
        'Upper Range Assigned for "other" EMEA sites',
        "PROD EMEA",
        "MGMT EMEA",
        "EMEA NON-AGENCY SPACE",
        "Administrative Container",
        "AGENCY EAST",
        "VOICE EMEA",
        "Administrative Container",
        "AGENCY WEST"})

    #  For filtering out data not needed.
    ddi_dict = {'MASTER': [],
                'Filt-Cidr-32': [],
                'Filt-100.88-Cidr-29': [],
                'Filt-100.64-Cidr-29': [],
                'Filt-Free-ip': [],
                'Filt-Cidr-15-to-Cidr-1': [],
                'Filt-Public-ip-View': [],
                'Filt-Wan_test-View': [],
                'Filt-OMC-IT-Parent-Subnet': [],
                'Filt-Leaf': [],
                'Filt-Dup': [],
                'Filt-Ignore': [],
                'Filt-Uncategorized': [],
                'Filt-Divest': [],
                'Full-Dataset': [],
                'Filt-Re-IP': [],
                'Filt-Drop Reserve': []}
    for i in range(rddifirst_sheet.nrows):
        if i == 0:
            continue
        ddi_dict['Full-Dataset'].append(rddifirst_sheet.row_values(i))
        if '/32' in rddifirst_sheet.row_values(i)[2]:
            ddi_dict['Filt-Cidr-32'].append(rddifirst_sheet.row_values(i))
            continue
        if 'leaf' in rddifirst_sheet.row_values(i)[20]:
            ddi_dict['Filt-Leaf'].append(rddifirst_sheet.row_values(i))
            continue
        if 'dup' in rddifirst_sheet.row_values(i)[20]:
            ddi_dict['Filt-Dup'].append(rddifirst_sheet.row_values(i))
            continue
        if 'ignore' in rddifirst_sheet.row_values(i)[20]:
            ddi_dict['Filt-Ignore'].append(rddifirst_sheet.row_values(i))
            continue
        if 're-ip' in rddifirst_sheet.row_values(i)[20]:
            ddi_dict['Filt-Re-IP'].append(rddifirst_sheet.row_values(i))
            continue
        if 'drop reserve' in rddifirst_sheet.row_values(i)[20]:
            ddi_dict['Filt-Drop Reserve'].append(rddifirst_sheet.row_values(i))
            continue
        if '100.88.0.0/29' in rddifirst_sheet.row_values(i)[3]:
            ddi_dict['Filt-100.88-Cidr-29'].\
                append(rddifirst_sheet.row_values(i))
            continue
        if '100.64.0.0/29' in rddifirst_sheet.row_values(i)[3]:
            ddi_dict['Filt-100.64-Cidr-29'].\
                append(rddifirst_sheet.row_values(i))
            continue
        if 'free ip' in rddifirst_sheet.row_values(i)[5].lower() or \
            'OPEN' in rddifirst_sheet.row_values(i)[5]:
                ddi_dict['Filt-Free-ip'].\
                    append(rddifirst_sheet.row_values(i))
                continue
        if rddifirst_sheet.row_values(i)[5] in omc_it_parent_list:
            ddi_dict['Filt-OMC-IT-Parent-Subnet'].append(
                rddifirst_sheet.row_values(i))
            continue
        if int(rddifirst_sheet.row_values(i)[2][1:3]) in range(1, 16):
            ddi_dict['Filt-Cidr-15-to-Cidr-1'].\
                append(rddifirst_sheet.row_values(i))
            continue
        if rddifirst_sheet.row_values(i)[4] == 'Public-IP':
            ddi_dict['Filt-Public-ip-View'].\
                append(rddifirst_sheet.row_values(i))
            continue
        if rddifirst_sheet.row_values(i)[20].strip() == 'divest':
            ddi_dict['Filt-Divest'].append(rddifirst_sheet.row_values(i))
            continue
        if rddifirst_sheet.row_values(i)[4] == 'wan_test':
            ddi_dict['Filt-Wan_test-View'].\
                append(rddifirst_sheet.row_values(i))
            continue
        if IPv4Network(rddifirst_sheet.row_values(i)[1]).is_private or \
                IPv4Network(rddifirst_sheet.row_values(i)[1]).is_cgn:
            ddi_dict['MASTER'].append(rddifirst_sheet.row_values(i))
        else:
            ddi_dict['Filt-Uncategorized'].append(rddifirst_sheet.
                                                  row_values(i))
    return ddi_dict


def _build_header_ea_index(header, e_att):
    with open(e_att, 'rb') as f_i:
        eatts = pickle.load(f_i)
    ea_idx_dict = {
        header[2]: eatts.index('Region_List'),
        header[3]: eatts.index('Country'),
        header[4]: eatts.index('City'),
        header[5]: eatts.index('Address'),
        header[6]: eatts.index('Site'),
        header[7]: eatts.index('Datacenter'),
        header[8]: eatts.index('Division'),
        header[9]: eatts.index('Requester Email'),
        header[10]: eatts.index('Agency'),
        header[11]: eatts.index('VLAN Description'),
        header[13]: eatts.index('Interface Name'),
        header[16]: eatts.index('IPR Designation')
        }
    return ea_idx_dict


def main():
    """
    Takes path and opens workbook.  Takes in DDI data and filters out
    unused data.  Once filtered it sends the array for writing.

    Output File:
       -- DDI_IPR_Sorted.xlsx
    """
    # get logger
    logger = logging.getLogger('ipr_format_ddi.py')
    logger.info('Beginning of Script')
    # Build paths and file names.
    raw_data_path = os.path.join(PROJECT_DIR, 'data', 'raw')
    interim_data_path = os.path.join(PROJECT_DIR, 'data', 'interim')
    processed_data_path = os.path.join(PROJECT_DIR, 'data', 'processed')
    ddi_file = os.path.join(raw_data_path, 'ddi_workbook.xls')
    ea_file = os.path.join(raw_data_path, 'ddi_dump_header.pkl')
    interim_unsorted_ddi_file = os.path.join(interim_data_path,
                                             'DDI_IPR_Unsorted.xlsx')
    processed_sorted_ddi_file = os.path.join(processed_data_path,
                                             'DDI_IPR_Sorted.xlsx')

    idx = _build_header_ea_index(HEADER_ROW, ea_file)

    logger.info('Filtering out unneeded data.')
    ddi_dict = _filter_data(ddi_file)
    logger.info('Filtering has been completed.')

    logger.info('Building Data Set for Sorting')
    _write_output_to_master(idx, ddi_dict, interim_unsorted_ddi_file)
    logger.info(r'Writing out DDI_IPR_Sorted.xlsx in data\processed.')
    _sorting_data(processed_sorted_ddi_file, interim_unsorted_ddi_file)
    logger.info('Script Complete')


if __name__ == '__main__':
    # getting root directory
    PROJECT_DIR = os.path.join(os.path.dirname(__file__), os.pardir, os.pardir)

    # setup logger
    LOG_FMT = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    logging.basicConfig(level=logging.INFO, format=LOG_FMT)

    # find .env automatically by walking up directories until it's found
    DOTENV_PATH = find_dotenv()

    # load up the entries as environment variables
    load_dotenv(DOTENV_PATH)

    # Headerrow for IPR Output
    HEADER_ROW = os.environ.get("IPR_HEADER_ROW").split(',')

    main()
