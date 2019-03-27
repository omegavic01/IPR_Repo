"""
This script takes the output from ipr_format_ddi.py.  It then performs
a validation check.  Once validation check passes it indexes and searches for
overlaps and conflict subnets.  Using the index number as the tag for either
overlap or conflict columns.

DataSets:
   DDI_to_IPR_Sorted.xlsx -- input file
   DDI-to-IPR.xlsx -- output file
"""
import logging
import os
import pickle
from collections import OrderedDict
from dotenv import find_dotenv, load_dotenv
from xlrd import open_workbook
from netaddr import IPNetwork
import openpyxl
from openpyxl.styles import Alignment
from checks.master_audit_validation_check import validation_check


def _wr_out_validation_check(master_sh, master_cidr_data, log_file):
    try:
        with open(log_file, mode='r') as file_input:
            adump = file_input.readlines()
        cidrcount = 0
        changes = False
        for item in adump:
            if 'Leading zero' in item:
                changes = True
                a_line = item.strip()
                blist = a_line.split(' ')
                newcidr = blist[-1]
                oldcidr = blist[0]
                for index, zero_data in enumerate(list(master_cidr_data)):
                    if oldcidr.strip() == zero_data[0].value.strip():
                        master_sh.cell(index+1, 2).value = newcidr
                        master_sh.cell(index+1, 18).value = int(newcidr
                                                                .split('.')[0])
                        master_sh.cell(index+1, 19).value = int(newcidr
                                                                .split('.')[1])
                        master_sh.cell(index+1, 20).value = int(newcidr
                                                                .split('.')[2])
                        master_sh.cell(index+1, 21).value = \
                            int(newcidr.split('.')[3].split('/')[0])
                        master_sh.cell(index+1, 22).value = \
                            int(newcidr.split('.')[3].split('/')[1])
            if 'Network is' in item:
                changes = True
                a_line = item.strip()
                blist = a_line.split(' ')
                newcidr = blist[-1]
                oldcidr = blist[0]
                for index, host_bits in enumerate(list(master_cidr_data)):
                    if oldcidr.strip() == host_bits[0].value.strip():
                        master_sh.cell(index+1, 2).value = newcidr
                        master_sh.cell(index+1, 18).value = int(newcidr
                                                                .split('.')[0])
                        master_sh.cell(index+1, 19).value = int(newcidr
                                                                .split('.')[1])
                        master_sh.cell(index+1, 20).value = int(newcidr
                                                                .split('.')[2])
                        master_sh.cell(index+1, 21).value = \
                            int(newcidr.split('.')[3].split('/')[0])
                        master_sh.cell(index+1, 22).value = \
                            int(newcidr.split('.')[3].split('/')[1])
            if 'out of range CIDR' in item:
                cidrcount += 1
    except FileNotFoundError:
        return 'Clean'
    if len(adump) == cidrcount:
        return 'Just Cidrs'
    if changes:
        return 'Changes'
    return 'Unclean'


def _pickle_data(master_ws, file):
    pickle_list = []
    for row in master_ws:
        templist = []
        for cell in row:
            templist.append(cell.value)
        pickle_list.append(templist)
    pickle.dump(pickle_list, open(file, 'wb'))


def _wr_out_overlap_conflict_tag(master_ws):
    master_rows = list(master_ws.rows)
    no_overlap_col = 26
    no_conflict_col = 27
    master_ws.cell(row=1, column=no_overlap_col,
                   value='No Overlap')
    master_ws.cell(row=1, column=no_conflict_col,
                   value='No Conflict')
    for enum, row in enumerate(master_rows):
        if enum == 0:
            continue
        if row[23].value is not None:
            master_ws.cell(row=enum + 1, column=no_overlap_col, value='NO')
        else:
            master_ws.cell(row=enum + 1, column=no_overlap_col, value='YES')
        if row[24].value is not None:
            master_ws.cell(row=enum + 1, column=no_conflict_col, value='NO')
        else:
            master_ws.cell(row=enum + 1, column=no_conflict_col, value='YES')


def _wr_out_conflict(master_ws, conflict_dicts, cidr_set):
    conflict_col = 25
    index_col = 22
    master_rows = list(master_ws.rows)
    master_ws.cell(row=1,
                   column=conflict_col,
                   value='Conflict Subnet - Index No.')
    for index, item in enumerate(cidr_set):
        if item in conflict_dicts:
            if len(conflict_dicts[item]) > 1:
                temp_list = conflict_dicts[item][:]
                if master_rows[index + 1][index_col].value in temp_list:
                    temp_list.remove(master_rows[index + 1][index_col].value)
                    if len(temp_list) == 1:
                        mycell = master_ws.cell(row=index + 2,
                                                column=conflict_col)
                        mycell.alignment = Alignment(horizontal='left')
                        mycell.value = int(temp_list[0])
                        continue
                    if len(temp_list) > 1:
                        master_ws.cell(row=index + 2, column=conflict_col,
                                       value=', '.join(str(x) for x in
                                                       temp_list))
                        continue


def _wr_out_overlap(master_ws, dict_of_overlaps, cidr_set):
    overlap_col = 24
    master_ws.cell(row=1,
                   column=overlap_col,
                   value='Conflict Subnet Overlap - Index No.')
    for index, item in enumerate(cidr_set):
        for key in dict_of_overlaps:
            if key != item:
                continue
            if not dict_of_overlaps[key] and key == item:
                continue
            if key == item and len(dict_of_overlaps[key]) == 1:
                mycell = master_ws.cell(row=index + 2, column=overlap_col)
                mycell.alignment = Alignment(horizontal='left')
                mycell.value = int(dict_of_overlaps[key][0])
                continue
            if key == item:
                master_ws.cell(row=index + 2, column=overlap_col,
                               value=', '.join(str(x) for x
                                               in dict_of_overlaps[key]))


def _conflict_overlap_check(interim_file):
    # Uses xlrd module for reading info
    master_wb = open_workbook(interim_file)
    m_sheet = master_wb.sheet_by_index(0)
    m_list_index = m_sheet.col_values(22)[1:]
    m_list_index = [int(x) for x in m_list_index]
    m_list_cidr = m_sheet.col_values(1)[1:]  # Ordered Cidr
    m_list_cidr_set = list(OrderedDict.fromkeys(m_list_cidr))
    m_cidr_index_zip = list(zip(m_list_cidr, m_list_index))
    m_dict_overlap = {}
    m_dict_conflict = {}
    for i in m_list_cidr_set:
        m_dict_overlap[i] = []
        m_dict_conflict[i] = []
    for item in m_cidr_index_zip:
        if item[0] in m_dict_conflict:  # Conflict check
            m_dict_conflict[item[0]].append(item[1])
    for item in m_cidr_index_zip:
        for _ip in m_list_cidr_set:
            if int(_ip.split('.')[0]) < int(item[0].split('.')[0]):
                continue
            if int(_ip.split('.')[1]) < int(item[0].split('.')[1]):
                continue
            elif _ip == item[0]:
                continue
            elif _ip.split('.')[0:2] == item[0].split('.')[0:2]:
                if IPNetwork(_ip) in IPNetwork(item[0]):
                    m_dict_overlap[_ip].append(item[1])
                    continue
            elif int(item[0].split('.')[1]) < int(_ip.split('.')[1]):
                break
            else:
                print(item)
    return m_dict_overlap, m_dict_conflict, m_list_cidr


def _indexing_data(interim_file, logger):
    master_wb = openpyxl.load_workbook(filename=interim_file)
    master_sheet = master_wb['MASTER']
    master_index_col = list(master_sheet.iter_cols(min_col=23, max_col=23))

    def _indexing(master_sh):
        ip_addr = 10002
        indexcol = 23
        master_view_col = list(master_sh.iter_cols(min_col=16, max_col=16))
        master_sh.cell(row=1, column=indexcol, value='Index')
        for index, item in enumerate(master_view_col[0]):
            if 'DDI View' in item.value:
                continue
            mycell = master_sh.cell(row=index + 1, column=indexcol)
            mycell.alignment = Alignment(horizontal='left')
            mycell.value = ip_addr
            ip_addr += 1

    if master_index_col[0][0].value is None:
        _indexing(master_sheet)
        master_wb.save(interim_file)
    else:
        logger.info('Indexing has been completed!')
        master_wb.save(interim_file)


def _ip_validation(processed_file, log_file, logger):
    master_wb = openpyxl.load_workbook(filename=processed_file)
    master_sheet = master_wb['MASTER']
    master_cidr_col = list(master_sheet.iter_rows(min_col=2, max_col=2))
    # IP Validation Check on Cidr Col.
    vcheck = validation_check(master_cidr_col, log_file)
    output_status = _wr_out_validation_check(master_sheet,
                                             master_cidr_col,
                                             log_file)
    master_wb.save(processed_file)
    if vcheck != 'Unclean':
        logger.info('Validation check was clean.')
    else:
        logger.info(r'Failed validation check: (Refer to) validation_log.txt')
        if output_status == 'Just Cidrs':
            logger.info(r'Just incorrectly assigned cidrs.  Continuing on.')
        if output_status == 'Changes':
            logger.info(r'Changes were made to the spreadsheet, '
                        r'please review log and spreadsheet.')
            exit()
        if output_status == 'Unclean':
            logger.info(r'Erred IPs are listed and need to be fixed.  '
                        r'Please review log and spreadsheet')
            exit()


def main():
    """
    Find and Notate Overlapping Subnets against Supernets

    Website for KB Purposes:

    https://stackoverflow.com/questions/35115138/how-do-i-check-if-
    a-network-is-contained-in-another-network-in-python

    Logic Identified:

    from netaddr import IPNetwork

    if IPNetwork("10.11.12.0/24") in IPNetwork("10.11.0.0/16"):
        print "Yes it is!"
    """
    logger = logging.getLogger('ipr_audit_ddi.py')
    logger.info('Beginning of Script')
    # Build path's.
    processed_data_path = os.path.join(PROJECT_DIR, 'data', 'processed')
    reports_data_path = os.path.join(PROJECT_DIR, 'reports')

    # Join file names to path's.
    interim_sorted_ddi_file = os.path.join(processed_data_path,
                                           'DDI_IPR_Sorted.xlsx')
    report_ddi_file = os.path.join(reports_data_path, 'DDI_to_IPR.xlsx')
    pickle_ddi_file = os.path.join(processed_data_path, 'ddi_to_ipr.pkl')
    validation_log_file = os.path.join(processed_data_path,
                                       'validation_log.txt')

    logger.info('Performing IP Validation Check')
    # IP Validation check:
    _ip_validation(interim_sorted_ddi_file, validation_log_file, logger)

    logger.info('Indexing Data')
    # Indexing for the networks listed.
    _indexing_data(interim_sorted_ddi_file, logger)

    logger.info('Performing Overlap and Conflict Check')
    # Overlap and Conflict check.
    dct_ovrlp, \
        dct_cnflct, \
        m_cidr_list = _conflict_overlap_check(interim_sorted_ddi_file)

    # Open workbook for final write of data.
    master_workbook = openpyxl.load_workbook(filename=interim_sorted_ddi_file)
    master_worksheet = master_workbook['MASTER']

    logger.info('Writing out compiled data.')
    # Write overlap data into final output.
    _wr_out_overlap(master_worksheet, dct_ovrlp, m_cidr_list)

    # Write conflict data into final output.
    _wr_out_conflict(master_worksheet, dct_cnflct, m_cidr_list)

    # Write in overlap and conflict tags.
    _wr_out_overlap_conflict_tag(master_worksheet)

    # Pickle data for future uses.
    _pickle_data(master_worksheet, pickle_ddi_file)

    # Save workbook.
    master_workbook.save(report_ddi_file)
    logger.info('Script Complete')


if __name__ == '__main__':
    # getting root directory
    PROJECT_DIR = os.path.join(os.path.dirname(__file__), os.pardir, os.pardir)

    # setup logger
    LOG_FMT = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    logging.basicConfig(level=logging.INFO, format=LOG_FMT)

    # find .env automatically by walking up directories until it's found
    DOTENV_PATH = find_dotenv()

    # load up the entries as environment variables
    load_dotenv(DOTENV_PATH)

    # Header Row for IPR Output
    HEADER_ROW = os.environ.get("IPR_HEADER_ROW").split(',')

    main()
