"""
This script is intended to take in the DDI data generated from the
ipam_query_app_full_report_xls.py script.  Twist, mash, split, and sort the
data.

DataSets:
   ddi_workbook.xls -- input file
   DDI_IPR_Unsorted.xlsx -- temp output file
   DDI_IPR_Sorted.xlsx -- output file
"""
import os
import logging
import operator
from xlrd import open_workbook
from ipaddr import IPv4Network
from openpyxl import Workbook
from openpyxl.styles import Alignment
from dotenv import find_dotenv, load_dotenv


def _sorting_data(interim_sorted_file, interim_unsorted_file):
    """
    Takes in the converted data from sort's all of the networks within
    each sheet within the work book.
    """
    input_wb = open_workbook(interim_unsorted_file)
    sheet_names = input_wb.sheet_names()
    output_wb = Workbook()
    for enum, sheet in enumerate(sheet_names):
        input_ws = input_wb.sheet_by_name(sheet)
        sorting_stuff = []  # Taking Unsorted data then sorting.
        for i in range(input_ws.nrows):
            if i == 0:
                continue
            sorting_stuff.append(input_ws.row_values(i))
        sorting_stuff = sorted(sorting_stuff,
                               key=operator.itemgetter(17, 18, 19, 20, 21))

        # Creating new spreadsheet with sorted data.

        def _write_final_output(sorted_stuff, sheet_name, idx):
            output_ws = output_wb.create_sheet(sheet_name, idx)
            output_ws.title = sheet_name
            row = 0
            for index, item in enumerate(HEADER_ROW):
                output_ws.cell(row=row + 1, column=index + 1, value=item)
            for stuff in sorted_stuff:
                row = row + 1
                for index, items in enumerate(stuff):
                    output_ws.cell(row=row + 1, column=index + 1, value=items)
        _write_final_output(sorting_stuff, sheet, enum)
    if 'Sheet' in output_wb.sheetnames:
        std = output_wb['Sheet']
        output_wb.remove(std)
    output_wb.save(interim_sorted_file)


def _write_output_to_master(ddi_dic, path):
    """
    Write out the rows in IPR expected format.

    Output Arguments:
        output_file - DDI-to-IPR-Format-Unsorted.xlsx
    """

    sheet_name_list = []
    for i in ddi_dic.keys():
        sheet_name_list.append(i)
    sheet_name_list.sort()
    sheet_name_list.reverse()

    # Buildout workbook!
    work_book = Workbook()
    for enum, sheet_name in enumerate(sheet_name_list):
        if not ddi_dic[sheet_name]:
            continue
        w_s = work_book.create_sheet(sheet_name, enum)  # work_sheet
        w_s.title = sheet_name
        row = 1
        col = 0
        for index, item in enumerate(HEADER_ROW):
            w_s.cell(row=row, column=index + 1, value=item)
        for bit in ddi_dic[sheet_name]:
            row = row + 1
            w_s.cell(row=row, column=col + 2, value=bit[3])  # Cidr
            w_s.cell(row=row, column=col + 3, value=bit[23])  # region
            w_s.cell(row=row, column=col + 4, value=bit[15])  # country
            w_s.cell(row=row, column=col + 5, value=bit[12])  # city
            w_s.cell(row=row, column=col + 6, value=bit[7])  # address
            w_s.cell(row=row, column=col + 7, value=bit[27])  # site
            w_s.cell(row=row, column=col + 8, value=bit[16])  # datacenter
            w_s.cell(row=row, column=col + 9, value=bit[17])  # division
            w_s.cell(row=row, column=col + 10, value=bit[26])  # email
            w_s.cell(row=row, column=col + 11, value=bit[8])  # agency
            w_s.cell(row=row, column=col + 13, value=bit[5])  # comment
            w_s.cell(row=row, column=col + 12, value=bit[31])  # vlandesc
            w_s.cell(row=row, column=col + 14, value=bit[19])  # interfacename
            w_s.cell(row=row, column=col + 15, value=bit[0])  # ddi_type
            w_s.cell(row=row, column=col + 16, value=bit[4])  # ddi_view
            w_s.cell(row=row, column=col + 17, value='DDI')
            mycell = w_s.cell(row=row, column=col + 18)
            mycell.alignment = Alignment(horizontal='left')
            mycell.value = int(bit[3].split('.')[0])  # 1st octet
            mycell = w_s.cell(row=row, column=col + 19)
            mycell.alignment = Alignment(horizontal='left')
            mycell.value = int(bit[3].split('.')[1])  # 2nd octet
            mycell = w_s.cell(row=row, column=col + 20)
            mycell.alignment = Alignment(horizontal='left')
            mycell.value = int(bit[3].split('.')[2])  # 3rd octet
            mycell = w_s.cell(row=row, column=col + 21)
            mycell.alignment = Alignment(horizontal='left')
            mycell.value = int(bit[3].split('.')[3].split('/')[0])  # 4th octet
            mycell = w_s.cell(row=row, column=col + 22)
            mycell.alignment = Alignment(horizontal='left')
            mycell.value = int(bit[3].split('/')[1])  # cidr for network addr.
    if 'Sheet' in work_book.sheetnames:
        std = work_book['Sheet']
        work_book.remove(std)
    work_book.save(path)


def main():
    """
    Takes path and opens workbook.  Takes in DDI data and filters out
    unused data.  Once filtered it sends the array for writing.

    Output Arguments:
        outputddilist = list of rows from spreadsheet.
    """
    # get logger
    logger = logging.getLogger('ddi_to_master.py')
    logger.info('Beginning of Script')
    # Build paths and file names.
    raw_data_path = os.path.join(PROJECT_DIR, 'data', 'raw')
    interim_data_path = os.path.join(PROJECT_DIR, 'data', 'interim')
    ddi_file = os.path.join(raw_data_path, 'ddi_workbook.xls')
    interim_unsorted_ddi_file = os.path.join(interim_data_path,
                                             'DDI_IPR_Unsorted.xlsx')
    interim_sorted_ddi_file = os.path.join(interim_data_path,
                                           'DDI_IPR_Sorted.xlsx')

    # Opens ddi_workbook.xls
    rddi = open_workbook(ddi_file)
    rddifirst_sheet = rddi.sheet_by_index(0)

    logger.info('Filtering out unneeded data.')
    #  For filtering out data not needed.
    ddi_dict = {'MASTER': [],
                'Filt-Cidr-32': [],
                'Filt-100.88-Cidr-29': [],
                'Filt-100.64-Cidr-29': [],
                'Filt-Free-ip-00890': [],
                'Filt-Cidr-15-to-Cidr-1': [],
                'Filt-Public-ip': [],
                'Filt-Wan_test': []}
    for i in range(rddifirst_sheet.nrows):
        if i == 0:
            continue
        if '/32' in rddifirst_sheet.row_values(i)[2]:
            ddi_dict['Filt-Cidr-32'].append(rddifirst_sheet.row_values(i))
            continue
        if '100.88.0.0/29' in rddifirst_sheet.row_values(i)[3]:
            ddi_dict['Filt-100.88-Cidr-29'].\
                append(rddifirst_sheet.row_values(i))
            continue
        if '100.64.0.0/29' in rddifirst_sheet.row_values(i)[3]:
            ddi_dict['Filt-100.64-Cidr-29'].\
                append(rddifirst_sheet.row_values(i))
            continue
        if 'free ip' in rddifirst_sheet.row_values(i)[5].lower() \
                and '00890' in rddifirst_sheet.row_values(i)[4]:
            ddi_dict['Filt-Free-ip-00890'].\
                append(rddifirst_sheet.row_values(i))
            continue
        if int(rddifirst_sheet.row_values(i)[2][1:3]) in range(1, 16):
            ddi_dict['Filt-Cidr-15-to-Cidr-1'].\
                append(rddifirst_sheet.row_values(i))
            continue
        if rddifirst_sheet.row_values(i)[4] == 'Public-IP':
            ddi_dict['Filt-Public-ip'].append(rddifirst_sheet.row_values(i))
            continue
        if rddifirst_sheet.row_values(i)[4] == 'wan_test':
            ddi_dict['Filt-Wan_test'].append(rddifirst_sheet.row_values(i))
            continue
        if IPv4Network(rddifirst_sheet.row_values(i)[1]).is_private:
            ddi_dict['MASTER'].append(rddifirst_sheet.row_values(i))
        elif IPv4Network(rddifirst_sheet.row_values(i)[1]).is_cgn:
            ddi_dict['MASTER'].append(rddifirst_sheet.row_values(i))

    # Send information for processing and to write output.
    logger.info('Building Data Set for Sorting')
    _write_output_to_master(ddi_dict, interim_unsorted_ddi_file)
    logger.info('Sorting Data and writing out DDI_IPR_Sorted.xlsx.')
    _sorting_data(interim_sorted_ddi_file, interim_unsorted_ddi_file)
    logger.info('Script Complete')


if __name__ == '__main__':
    # getting root directory
    PROJECT_DIR = os.path.join(os.path.dirname(__file__), os.pardir, os.pardir)

    # setup logger
    LOG_FMT = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    logging.basicConfig(level=logging.INFO, format=LOG_FMT)

    # find .env automatically by walking up directories until it's found
    DOTENV_PATH = find_dotenv()

    # load up the entries as environment variables
    load_dotenv(DOTENV_PATH)

    # Headerrow for IPR Output
    HEADER_ROW = os.environ.get("IPR_HEADER_ROW").split(',')

    main()
